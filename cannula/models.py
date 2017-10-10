from django.db import models
from django.db.models.signals import post_init
from django.core.files.storage import FileSystemStorage
from django.conf import settings

import logging
logger = logging.getLogger(__name__)

import mimetypes
from functools import lru_cache, partialmethod, partial

from . import grabbag

def make_random_filename(instance, filename):
    mt = mimetypes.guess_type(filename)
    file_ext = mimetypes.guess_extension(mt[0])

    return grabbag.make_random_code(code_length=16) + file_ext

fs = FileSystemStorage(location=settings.SOURCE_DOC_DIR)

class SourceDocument(models.Model):
    orig_filename = models.CharField(max_length=128, blank=True, null=True)
    file = models.FileField(upload_to=make_random_filename, storage=fs)
    uploaded_at = models.DateTimeField(auto_now_add=True)

    def save(self, *args, **kwargs):
        # store the original filename away for later
        self.orig_filename = self.file.name
        super(SourceDocument, self).save(*args, **kwargs)

    def __str__(self):
        return '%s: %s' % (self.file, self.orig_filename)

class DataElement(models.Model):
    VALUE_TYPES = (
        ('NUMBER', 'Number'),
        ('INTEGER', 'Integer (whole numbers only)'),
        ('POS_INT', 'Positive Integer'),
        # ('PERCENT', 'Percentage'), # implies the Average aggregation method
        #TODO: for boolean types auto-create a (hidden?) category with two options corresponding to the labels we have in the boolean
        # ('BOOLEAN', 'Boolean (True/False)'), # two values (0/1) with default labels (category options?) of 'True'/'False'
        # ('CHOICE', 'Selection from fixed set') # implies the Count aggregation method
    )
    AGG_METHODS = (
        ('SUM', 'Sum()'),
        # ('COUNT', 'Count()'), # eg. we have a facility reporting patient aggregates but also a facility level field (has_incinerator)
        # ('AVG', 'Average()'), # needs a corresponding 'population' data element that provides the ratios when we combine two (or more) averages
    )
    name = models.CharField(max_length=128)
    value_type = models.CharField(max_length=8, choices=VALUE_TYPES)
    value_min = models.DecimalField(max_digits=17, decimal_places=4, verbose_name='Minimum Value', blank=True, null=True)
    value_max = models.DecimalField(max_digits=17, decimal_places=4, verbose_name='Maximum Value', blank=True, null=True)
    aggregation_method = models.CharField(max_length=8, choices=AGG_METHODS)

    def __repr__(self):
        return 'DataElement<%s>' % (str(self),)

    def __str__(self):
        return '%s' % (self.name,)


# TODO: Consider tracking which data element each subcategory is from (reduce false matches and other? benefits)
CATEGORIES = [
    'Male',
    'Female',
    '18 Mths-<5 Years',
    '5-<10 Years',
    '10-<15 Years',
    '15-<19 Years',
    '19-<49 Years',
    '>49 Years',
    '10-19 Years',
    '20-24 Years',
    '>=25 Years',

    '<2 Years',
    '2 - < 5 Years (HIV Care)',
    '5 - 14 Years',
    '< 15 Years',
    '15 Years and above',
    # HMIS 106a: 1B ART QUARTERLY COHORT ANALYSIS REPORT: FOLLOW-UP
    'Alive on ART in Cohort',
    'Died',
    'Lost  to Followup',
    'Lost',
    'Started on ART-Cohort',
    'Stopped',
    'Transfered In',
    'Transferred Out',
    # HMIS 105: 1.3 OPD
    '0-28 Days',
    '29 Days-4 Years',
    '5-59 Years',
    '60andAbove Years',
]

import re
SEP_REGEX = '[\s,]+' # one or more of these characters in sequence
CATEGORY_REGEX = '|'.join('%s?(%s)' % (SEP_REGEX, re.escape(categ)) for categ in CATEGORIES)

def unpack_data_element(de_long):
    m = re.split(CATEGORY_REGEX, de_long)
    # squash list of matches by removing blank and None entries (and False and numeric zeroes)
    de_name, *category_list = tuple(filter(None, m))
    cat_str = ', '.join(category_list)

    # deals with cases where the data element name includes a subcategory ('105-2.1a Male partners received HIV test results in eMTCT')
    # and matches multiple subcategories ('Lost' and 'Lost  to Followup' in '106a Cohort  All patients 12 months Lost  to Followup')
    #TODO: reimplement this, it is a really ugly hack
    if any([cat not in CATEGORIES for cat in category_list]):
        cat_str = ' '.join(category_list)
        if cat_str not in CATEGORIES:
            de_name = de_long
            cat_str = ''

    de_instance, created = DataElement.objects.get_or_create(name=de_name, value_type='NUMBER', value_min=None, value_max=None, aggregation_method='SUM')
    return (de_instance, cat_str)

class DataValue(models.Model):
    data_element = models.ForeignKey(DataElement, related_name='data_values')
    #TODO: break this out into a foreign key to the Category/Subcategory models
    category_str = models.CharField(max_length=128)
    site_str = models.CharField(max_length=128)
    numeric_value = models.DecimalField(max_digits=17, decimal_places=4)
    month = models.CharField(max_length=7, blank=True, null=True) # ISO 8601 format '2017-09'
    quarter = models.CharField(max_length=7, blank=True, null=True) # ISO 8601 format '2017-Q3'
    year = models.CharField(max_length=4, blank=True, null=True) # ISO 8601 format '2017'
    source_doc = models.ForeignKey(SourceDocument, related_name='data_values')

    def __repr__(self):
        return 'DataValue<%s [%s], %s, %s, %d>' % (str(self.data_element), self.category_str, self.site_str,  next(filter(None, (self.month, self.quarter, self.year))), self.numeric_value,)

    def __str__(self):
        return '%s [%s], %s, %s, %d' % (str(self.data_element), self.category_str, self.site_str.split(' => ')[-1],  next(filter(None, (self.month, self.quarter, self.year))), self.numeric_value,)

@lru_cache(maxsize=16) # memoize to reduce cost of "parsing"
def extract_periods(period_str):
    from .grabbag import period_to_dates, dates_to_iso_periods
    dates = period_to_dates(period_str)
    return dates_to_iso_periods(*dates)

def load_excel_to_datavalues(source_doc, max_sheets=4):
    from collections import defaultdict
    from datetime import datetime
    from itertools import islice
    import re
    import calendar
    import openpyxl

    MONTH_REGEX = r'[\s]*(%s) [0-9]{4}[\s]*' % ('|'.join(calendar.month_name[1:]),)
    MONTH_PREFIX_REGEX = r'^[\s]*(%s) ([0-9]{4})?[\s]*' % ('|'.join(calendar.month_name[1:]),)

    DE_COLUMN_START = 4 # 0-based index of first dataelement column in worksheet

    wb = openpyxl.load_workbook(source_doc.file.path)
    logger.debug(wb.get_sheet_names())

    wb_loc_values = defaultdict(list) # when a new key is encountered return a new empty list

    for ws_name in ['Step1']: #wb.get_sheet_names()[:max_sheets]:
        ws = wb[ws_name]
        logger.debug((ws_name, ws.max_row, ws.max_column))

        headers = [cell.value for cell in ws.rows[0]]
        # discard the month (and space) prefix on the data element names
        clean_headers = (re.sub(MONTH_PREFIX_REGEX, '', h) for h in headers[DE_COLUMN_START:] if h is not None)
        data_elements = (*(unpack_data_element(de) for de in clean_headers),)
        

        for row in ws.rows[1:]: # skip header row
            period, *location_parts = [c.value for c in row[:DE_COLUMN_START]]
            if not period:
                continue
            iso_year, iso_quarter, iso_month = extract_periods(period.strip())
            location = ' => '.join(location_parts)
            logger.debug((period, location))

            site_val_cells = row[DE_COLUMN_START:]
            site_values = zip(data_elements, (c.value for c in site_val_cells))
            dv_construct = partial(DataValue, site_str=location, month=iso_month, quarter=iso_quarter, year=iso_year, source_doc=source_doc)
            data_values = [dv_construct(data_element=sv[0][0], category_str=sv[0][1], numeric_value=sv[-1]) for sv in site_values if sv[-1] != '' and not(sv[-1] is None)]
            
            wb_loc_values[location] += data_values

    return dict(wb_loc_values) # convert back to a normal dict for our callers
